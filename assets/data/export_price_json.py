from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook


DATA_DIR = Path(__file__).resolve().parent
WORKBOOK_PATH = DATA_DIR / "price.xlsx"
OUTPUT_ROOT = DATA_DIR
STYLES_DIR = OUTPUT_ROOT / "styles"

# price.xlsx 目前只有 style 工作表，因此 model / color 先由這裡統一維護。
DEFAULT_MODELS = ["1027", "1027S", "827", "1027T"]
DEFAULT_COLORS = ["311", "8195", "49316", "ST"]

STYLE_SHEETS = [
	"2D",
	"F_2D",
	"2D_2D",
	"3D",
	"F_3D",
	"4D",
	"F_4D",
	"4D_4D",
	"6D",
	"F_6D",
]

# Excel 排版規格：
# - 第 3 列：B 欄為 H/W，C 欄起為 widths，讀到第一個空白欄為止
# - 第 4 列起：B 欄為 height，C 欄起為對應價格，寬度數量與第 3 列一致
WIDTH_HEADER_ROW = 3
WIDTH_START_COLUMN = 3
HEIGHT_START_ROW = 4
HEIGHT_COLUMN = 2


def round_price(value: float | int | None) -> int:
	if value is None:
		raise ValueError("Price cell is empty.")
	return int(round(float(value)))


def extract_widths(worksheet) -> list[int]:
	widths: list[int] = []
	column = WIDTH_START_COLUMN

	while True:
		value = worksheet.cell(row=WIDTH_HEADER_ROW, column=column).value
		if value is None or str(value).strip() == "":
			break
		widths.append(int(value))
		column += 1

	if not widths:
		raise ValueError(f"{worksheet.title}: no widths found in row {WIDTH_HEADER_ROW}.")

	return widths


def extract_heights_and_values(worksheet, width_count: int) -> tuple[list[int], list[list[int]]]:
	heights: list[int] = []
	values: list[list[int]] = []
	row = HEIGHT_START_ROW

	while True:
		height_value = worksheet.cell(row=row, column=HEIGHT_COLUMN).value
		if height_value is None or str(height_value).strip() == "":
			break
		if not isinstance(height_value, (int, float)):
			break

		heights.append(int(height_value))
		row_values: list[int] = []

		for index in range(width_count):
			price_value = worksheet.cell(row=row, column=WIDTH_START_COLUMN + index).value
			row_values.append(round_price(price_value))

		values.append(row_values)
		row += 1

	if not heights:
		raise ValueError(f"{worksheet.title}: no heights found from row {HEIGHT_START_ROW}.")

	return heights, values


def export_style_sheet(worksheet) -> dict:
	widths = extract_widths(worksheet)
	heights, values = extract_heights_and_values(worksheet, len(widths))

	return {
		"style": worksheet.title,
		"widths": widths,
		"heights": heights,
		"values": values,
	}


def write_json(path: Path, payload) -> None:
	path.parent.mkdir(parents=True, exist_ok=True)
	path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def main() -> None:
	if not WORKBOOK_PATH.exists():
		raise SystemExit(f"Workbook not found: {WORKBOOK_PATH}")

	workbook = load_workbook(WORKBOOK_PATH, data_only=True)

	write_json(OUTPUT_ROOT / "models.json", DEFAULT_MODELS)
	write_json(OUTPUT_ROOT / "colors.json", DEFAULT_COLORS)

	exported_styles: list[str] = []

	for style_name in STYLE_SHEETS:
		if style_name not in workbook.sheetnames:
			raise SystemExit(f"Missing worksheet: {style_name}")

		payload = export_style_sheet(workbook[style_name])
		write_json(STYLES_DIR / f"{style_name}.json", payload)
		exported_styles.append(style_name)

	print("Export complete")
	print(f"Workbook: {WORKBOOK_PATH}")
	print(f"Output root: {OUTPUT_ROOT}")
	print(f"Styles dir: {STYLES_DIR}")
	print(f"Models: {len(DEFAULT_MODELS)}")
	print(f"Colors: {len(DEFAULT_COLORS)}")
	print(f"Styles: {', '.join(exported_styles)}")


if __name__ == "__main__":
	main()
